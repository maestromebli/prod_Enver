#!/usr/bin/env python3
"""Парсинг замовлень з Excel: колонки A+B та додаткові рядки в колонці I (номер + назва в одній комірці)."""
import json
import re
import sys
from openpyxl import load_workbook

ORDER_RE = re.compile(r"^\(?(Е[МM]?-\d+(?:/\d+)?)\)?\s+(.+)$", re.IGNORECASE)
SKIP_VALUES = {"Об`єкт:", "%", "№"}


def norm_num(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in SKIP_VALUES:
        return None
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1].strip()
    return s.strip()


def add_order(orders, seen, order_number, object_name):
    n = norm_num(order_number) or str(order_number).strip()
    obj = str(object_name).strip() if object_name is not None else ""
    if not n or not obj:
        return
    if n in seen:
        return
    seen.add(n)
    orders.append({"orderNumber": n, "object": obj})


def parse_combined_cell(value):
    s = str(value).strip()
    if not s or s in SKIP_VALUES:
        return None, None
    m = ORDER_RE.match(s)
    if m:
        return m.group(1), m.group(2).strip()
    return None, None


def main():
    path = sys.argv[1]
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    orders = []
    seen = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        row = list(row) + [None] * 12
        n, name = norm_num(row[0]), row[1]
        if n and name is not None and str(name).strip():
            add_order(orders, seen, n, name)
        extra_num, extra_name = parse_combined_cell(row[8])
        if extra_num and extra_name:
            add_order(orders, seen, extra_num, extra_name)
    wb.close()
    print(json.dumps(orders, ensure_ascii=False))


if __name__ == "__main__":
    main()
